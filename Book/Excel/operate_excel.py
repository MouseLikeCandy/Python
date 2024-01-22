"""
xlrd    读取工作簿     *.xls/*.xlsx
xlwt    写入工作簿     *.xls

EXCEL:
*.xls 2003
*.xlsx 2007
工作簿 -> 工作表 -> 单元格
"""
import xlrd
import xlwt
from faker import Faker                 # 从faker模块导入Faker这个类
from openpyxl import load_workbook
import pandas

# 实例化，保存到变量fake中
fake = Faker(locale='zh_CN')

# people = xlrd.open_workbook('people.xlsx')
#
# sheet = people.sheets()[0]
# sheet = people.sheet_by_index(0)
# sheer = people.sheet_by_name('Sheet1')

# wb = load_workbook("1.xlsx")
#
# for sheetName in wb.sheetnames:
#     sheet = wb[sheetName]
#     for rownum in range(2, 100):
#         sheet.cell(row=rownum, column=1).value = fake.name()
#         sheet.cell(row=rownum, column=2).value = fake.country()
#         sheet.cell(row=rownum, column=3).value = fake.province()
#         sheet.cell(row=rownum, column=4).value = fake.address()
#         sheet.cell(row=rownum, column=5).value = fake.company()
#         sheet.cell(row=rownum, column=6).value = fake.postcode()
#         sheet.cell(row=rownum, column=7).value = fake.ssn(min_age=18, max_age=58)
#         sheet.cell(row=rownum, column=8).value = fake.fake.date(pattern="%Y-%m-%d", end_datetime=None)
#         sheet.cell(row=rownum, column=9).value = fake.phone_number()
#         sheet.cell(row=rownum, column=10).value = fake.email()
#         sheet.cell(row=rownum, column=11).value = fake.credit_card_number(card_type=None)





