# -*- coding: UTF-8 -*-
"""
@Project ：Python 
@File    ：excel_openpyxl.py
@IDE     ：PyCharm 
@Author  ：Ning
@Date    ：2023/5/9 13:35 
"""
import openpyxl

"""
openpyxl
处理大型工作簿文件
同时支持*.xls和*.xlsx

劣势:对Excel中的VBA支持并不友好
VBA远没有Python强大和易用, 掌握Python, 不再使用VBA

注意:openpyxl与xlrd/xlwt不同,cell方法中的row/column参数从1开始
"""
'''
# 读取
'''

# 打开已有的*.xlsx文件
work_book = openpyxl.load_workbook('people_openpyxl.xlsx')
# 选择第一个工作表
work_sheet = work_book.worksheets[0]
# 获取第3行第2列的值
work_sheet.cell(row=3, column=2).value


# 获取工作表的其他数据
# 返回sheet中有数据的最大行数
print('最大行数:', work_sheet.max_row)
print('最小行数:', work_sheet.min_row)
print('最大列数:', work_sheet.max_column)
print('最小列数:', work_sheet.min_column)

# 循环处理工作表
for col in work_sheet.iter_cols(min_col=3, max_col=5, max_row=2, values_only=True):
    print(col)

# 获取工作表中所有的值,只读模式下不可用
all_values = work_sheet.values  # 返回一个生成器对象
print(type(all_values))
for i, value in enumerate(all_values):
    print(value)
    if i == 3:
        break

'''
# 写入
'''
# 创建工作簿对象
wb = openpyxl.Workbook()
# wb.actice 默认返回第一个工作表
ws = wb.active
# 第一个工作表名称
print('ws title:', ws.title)
# 创建一个新的worksheet
ws2 = wb.create_sheet("NewTitle", 1)
# 修改Title
ws2.title = 'MySheet'
# 添加内容
ws2.cell(row=2, column=2).value = '二两'
# 保存
wb.save('openpyxl_test.xlsx')

'''
# 修改内容
'''
wb = openpyxl.load_workbook('openpyxl_test.xlsx')
ws = wb.active
# 修改A2~C3区域的值
for row in ws['A2':'C3']:
    for cell in row:
        cell.value = 'new value'
# 保存
wb.save('openpyxl_test.xlsx')


'''
# 修改样式
'''
# 创建工作簿对象
wb = openpyxl.Workbook()
ws = wb.active
rows = [
    ['ID', 'name', 'age'],
    [1, '张三', 28],
    [2, '李四', 25],
    [3, '王五', 40],
    [4, '赵六', 23]
]

for row in rows:
    # 添加多行
    ws.append(row)

from openpyxl.styles import Font, colors

# 字体设置为微软雅黑, 字体大小25,斜体,红色
font = Font(name='微软雅黑', size=25, italic=True, color='FF0000', bold=True)
# 设置对应单元格的字体样式
ws['A1'].font = font

from openpyxl.styles import PatternFill
# 填充样式, 将单元格背景色填充为绿色
fill = PatternFill(fill_type='solid', start_color=colors.BLUE)
ws['B1'].fill = fill

from openpyxl.styles import Border,Side
# 边框样式
border = Border(
    left=Side(border_style='double', color='FFBB00'),
    right=Side(border_style='double', color='FFBB00'),
    top=Side(border_style='double', color='FFBB00'),
    bottom=Side(border_style='double', color='FFBB00'),
)
ws['C1'].border = border

from openpyxl.styles import Alignment
# 单元格内容对齐
align = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws['D1'].alignment = align

# 修改行高列宽
# 第3行行高修改为40
ws.row_dimensions[3].height = 40
# A列列宽修改为30
ws.column_dimensions['A'].width = 30

# 合并一行中的几个单元格
ws.merge_cells('A7:C7')
# 合并一个矩形区域中的单元格
ws.merge_cells('A9:C13')
ws['A9'] = '合并单元格'

wb.save('openpyxl_styles.xlsx')

'''
# 使用openpyxl操作大型工作簿时
# read_only模式, write_only模式
'''

from openpyxl import load_workbook
# 在read_only模式下载入大型工作簿big.xlsx
wb = load_workbook(filename='big.xlsx', read_only=True)
# 选择big_sheet工作表
ws = wb['big_sheet']

for row in ws.rows:
    for cell in row:
        print(cell.value)

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.comments import Comment
from openpyxl.styles import Font

# write_only设置为True
wb = Workbook(write_only=True)
# write_only模式下,不会包含任何工作表,需要使用create_sheet方法自行创建
ws = wb.create_sheet()
# write_only模式下,单元格想要具有样式,就只能使用WriteOnlyCell创建单元格
cell = WriteOnlyCell(ws, value='write_only状态写入的内容')
# 为单元格设置字体样式
cell.font = Font(name='微软雅黑', size=36)
# 插入Excel批注
cell.comment = Comment(text='这是一个批注', author='二两')
# write_only模式下只能使用append方法添加数据
ws.append([cell, 2.333, None])
# 保存
wb.save('write_only.xlsx')

