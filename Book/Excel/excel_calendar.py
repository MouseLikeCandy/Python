# -*- coding: UTF-8 -*-
"""
@Project ：Python 
@File    ：excel_calendar.py
@IDE     ：PyCharm 
@Author  ：Ning
@Date    ：2023/5/11 8:54 
"""


'''
openpyxl
'''

'''1. 获取日历数据'''
import calendar
# 指定一周的第一天, 0是星期一(默认值), 6是星期天
calendar.setfirstweekday(firstweekday=6)

year = 2023
# 循环月份
for i in range(1, 13):
    # 每月中的每一行, 一行表示一周
    for j in range(len(calendar.monthcalendar(year, i))):
        # 每一天
        for k in range(len(calendar.monthcalendar(year, i)[j])):
            # 具体的日期
            value = calendar.monthcalendar(year, i)[j][k]

'''2. 将日历数据写入工作表'''
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font
import calendar
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image


# 创建工作簿
wb = openpyxl.Workbook()
year = 2023
# for i in range(12, 0, -1):
for i in range(1, 13):
    # 添加工作表,每个月份对应一个工作表
    sheet = wb.create_sheet(index=0, title=str(i) + '月')
    for j in range(len(calendar.monthcalendar(year, i))):
        # 每一天
        for k in range(len(calendar.monthcalendar(year, i)[j])):
            # 具体的日期
            value = calendar.monthcalendar(year, i)[j][k]
            if value == 0:
                value = ''  # 将0值变为空值, 没有日期的单元格填空值
                sheet.cell(row=j+9, column=k+1).value = value
            else:
                # 将日期数据添加到具体的单元格中
                sheet.cell(row=j+9, column=k+1).value = value
                # 设置字体
                sheet.cell(row=j+9, column=k+1).font = Font(u'微软雅黑', size=11)
    '''3. 修改单元格中文字的位置'''
    # 单元格文字设置, 右对齐, 垂直居中
    align = Alignment(horizontal='right', vertical='center')
    # 单元格填充色属性设置
    fill = PatternFill('solid', fgColor='99CCCC')
    # 对单元格进行颜色填充
    for k1 in range(1, 50):
        for k2 in range(1, 50):
            sheet.cell(row=k1, column=k2).fill = fill
    '''4. 日期对应星期几'''
    # 星期日开头, 符合calendar.setfirstweekday(firstday=6)设置
    days = ['星期日', '星期一', '星期二', '星期三', '星期四', '星期五', '星期六']
    num = 0
    # 添加星期几相关信息
    for k3 in range(1, 8):
        sheet.cell(row=8, column=k3).value = days[num]
        # 设置样式
        sheet.cell(row=8, column=k3).alignment = align
        sheet.cell(row=8, column=k3).font = Font(u'微软雅黑', size=11)
        # 设置列宽12
        c_char = get_column_letter(k3)
        sheet.column_dimensions[c_char].width = 12
        num += 1
    # 将日历所在单元格的行高都修改为30
    for k4 in range(8, 14):
        sheet.row_dimensions[k4].height = 30

    '''5. 插入图片, 使用openpyxl库中的Image类, 该类依赖Pillow图像处理库'''
    # pip install Pillow
    # 合并单元格
    sheet.merge_cells('I1:P20')
    # 添加图片
    img = Image('1.png')
    # 设置图片大小
    newsize = (200, 200)
    img.width, img.height = newsize
    # 与顶部有些距离, 好看一些, 顶部为I1
    sheet.add_image(img, 'I2')

    '''6. 添加年月文字'''
    # 添加年份月份
    sheet.cell(row=3, column=1).value = f'{year}年'
    sheet.cell(row=4, column=1).value = str(i) + '月'
    # 设置年份月份文本样式
    sheet.cell(row=3, column=1).font = Font(u'微软雅黑', size=16, bold=True, color='FF7887')
    sheet.cell(row=4, column=1).font = Font(u'微软雅黑', size=16, bold=True, color='FF7887')
    # 设置年份月份文本对齐方式
    sheet.cell(row=3, column=1).alignment = align
    sheet.cell(row=4, column=1).alignment = align

# 保存
wb.save('calendar.xlsx')