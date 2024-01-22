"""
pip install xlrd        # 读取    xls xlsx
pip install xlwt        # 写入    xls
pip install xlutils     # 添加工作表
pip install openpyxl    # 添加工作表
pip install pandas

*.xls  2003   复合文档类型结构
*.xlsx 2007   XML数据结构

pip install datetime -i http://pypi.douban.com/simple --trusted-host pypi.douban.com
"""
import openpyxl
from faker import Faker
from openpyxl import Workbook
import xlrd
from xlutils.copy import copy as xl_copy
import datetime

# 创建faker实例,指定语言(区域)
faker = Faker(locale="zh_CN")  # 中文
# 调用方法来获取模拟数据
# name()    模拟名字
# email()   模拟邮箱
# phone_number() 模拟手机号
# address() 模拟地址
# print(faker.name(), faker.email(), faker.phone_number(), faker.address(), faker.company())

# #创建工作表和工作簿
wb1 = Workbook()
active_sheet = wb1.active
active_sheet.append(['姓名', '邮箱', '电话', '地址', '公司'])
print(f'active_sheet: {active_sheet}, type: {type(active_sheet)}')
# # 追加数据
# sheet.append()追加一行数据
# append([])
for i in range(100):
    active_sheet.append([faker.name(), faker.email(), faker.phone_number(),
                         faker.address(), faker.company()])
wb1.save("faker_new.xlsx")

# # 向已存在的Excel中添加新的工作表
# 1.使用xlutils.copy
# open existing workbook # 报错解决：在桌面上手动创建Excel文件即可解决。
rb = xlrd.open_workbook('faker2.xls')
# make a copy of it
wb2 = xl_copy(rb)
# add sheet to workbook with existing sheets
now_time = datetime.datetime.now().strftime('%F %H %M %S')
Sheet1 = wb2.add_sheet(f'{now_time}')  # 不允许重复的工作表

#  方法二：使用openpyxl库
wb3 = openpyxl.load_workbook(r'faker.xlsx')
# 连续三个Sheet2, 创建3张表Sheet2,Sheet21,Sheet22
wb3.create_sheet(title=f'{now_time}')  # 允许重复，如果有数据，会被清空
# wb3.create_sheet(title='Sheet2')
# wb3.create_sheet(title='Sheet2')
# wb3.create_sheet(title='Sheet2', index=0)

wb2.save('faker2.xls')
wb3.save(r'faker.xlsx')

# # import xlrd
faker_people = xlrd.open_workbook('faker.xlsx')
'''
PermissionError: [Errno 13] Permission denied: 'people.xlsx'
拒绝访问  - 权限  - 打开
'''
'''
xlrd.biffh.XLRDError: Excel xlsx file; not supported
最新的xlrd居然不支持Excel xlsx文件的读取.方法：卸载最新的xlrd库，
安装历史版本的xlrd库。pip install xlrd ==版本号
pip install xlrd==1.2.0
查看版本：
pip show xlrd
pip list
卸载：
pip uninstall xlrd
pip uninstall xlrd -y
'''

sheets = faker_people.sheets()
sheet1 = faker_people.sheets()[0]
print(f'sheet1 is {sheet1}')
sheet2 = faker_people.sheet_by_index(0)
print(f'sheet2 is {sheet2}')
sheet3 = faker_people.sheet_by_name('Sheet')
print(f'sheet3 is {sheet3}')

sheets_num = faker_people.nsheets
sheets_names = faker_people.sheet_names()
n_rows = sheet3.nrows
n_cols = sheet3.ncols
print(f'工作表数量：{sheets_num}，工作表名称列表：{sheets_names}, 行数：{n_rows}, 列数：{n_cols}')

'''
DeprecationWarning: Call to deprecated function get_sheet_by_name (Use wb[sheetname]).
不推荐警告：调用已弃用的函数get_sheet_by_name（请使用wb[sheetname])
'''
# # # 继续向表中加入假数据
for sheet_name in sheets_names:
    cur_sheet = wb3[sheet_name]
    print(f'cur_sheet: {cur_sheet}, type: {type(cur_sheet)}')
    cur_sheet.append(['姓名', '邮箱', '电话', '地址', '公司'])
    # # 追加数据
    # sheet.append()追加一行数据
    # append([])
    for i in range(100):
        cur_sheet.append([faker.name(), faker.email(), faker.phone_number(),
                          faker.address(), faker.company()])

wb3.save(r'faker.xlsx')


for sheet in sheets:
    n_rows = sheet.nrows
    n_cols = sheet.ncols
    print(f'当前表：{sheet}')
    for row in range(1, n_rows):
        for col in range(n_cols):
            cell_value = sheet3.cell_value(row, col)
            print(cell_value, end=' ')
        print('--' * 50)
    print('/' * 50)

